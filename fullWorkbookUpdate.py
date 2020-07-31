import openpyxl
import sys
import os.path

excelFileName = sys.argv[1]
scriptName = sys.argv[2]
functionalArea = sys.argv[3]

fileLocation = '/Users/michaelkirschbaum/Desktop/Python/excelToSQL/'

##
# Creates 2d List
#
def setUpList(data):
    masterList = []
    for i in data:
        masterList.append(list(i))
    return masterList
#
# Helper Function not really needed
#
def prettyPrint(myList):
    for i in range(len(myList)):
        for j in range(len(myList[i])):
            print(str(myList[i][j]), '\t', end =" ")
        print()

#
# Gets table row names, for query later
#
def createRowNames(line):
    line = ", ".join(line)
    return line

#
# Formats the numbers and strings of the values correctly
#
def formatValues(myList):
    #Formats the correct
    for i in range(1, len(myList)):
        if 'collection_threshold_amount' in myList[0]:
            collectionThresholdAmount = myList[0].index('collection_threshold_amount')
            myList[i][collectionThresholdAmount] = '{:.4f}'.format(float(myList[i][collectionThresholdAmount].encode('ascii', 'ignore')))
        if 'refund_fraud_threshold_amt' in myList[0]:
            refundFraudThresholdAmt = myList[0].index('refund_fraud_threshold_amt')
            myList[i][refundFraudThresholdAmt] = '{:.4f}'.format(float(myList[i][refundFraudThresholdAmt].encode('ascii', 'ignore')))
        for j in range(len(myList[i])):
            if isinstance(myList[i][j], str) and myList[i][j] != 'GETDATE()' and myList[i][j] != 'NULL':
                myList[i][j] = "'" + myList[i][j]+"'"
    return myList

#
# Creates one full query
#
def createQuery(valueList, rowNames, sheetname):
    if 'code' in rowNames:
        codeString = rowNames
        codeString = codeString.split('code', 1)[0]
        counter = codeString.count(',')
        whereCount = counter
        whereValue = 'code = ' + valueList[whereCount]
    elif 'desc_short' in rowNames:
        desc_shortString = rowNames
        desc_shortString = desc_shortString.split('desc_short', 1)[0]
        counter = desc_shortString.count(',')
        whereCount = counter
        whereValue = 'desc_short = ' + valueList[whereCount]
    elif '_x_' in rowNames:
        crossRefString = rowNames
        crossRefString = crossRefString.split(',', 1)[1]
        crossRefStringList = crossRefString.split(',')
        removeList = []
        for i in crossRefStringList:
            if '_key' not in i:
                removeList.append(i)
        for i in removeList:
            crossRefStringList.remove(i)
        whereCount = len(crossRefStringList)
        whereValueList = valueList[1:]
        whereValueList = whereValueList[0:whereCount]
        whereValue = dict(zip(crossRefStringList,whereValueList))
        whereValue = str(whereValue).replace("'",'').replace('{','').replace('}','').replace(':',' =').replace('  ', ' ').replace('"',"'").replace(',',' and')
    else:
        uniqueCount = 1
        whereValue = 'core = ' + valueList[uniqueCount]

    whereStatement = 'WHERE {}'.format(whereValue)
    setList = rowNames.split(',')
    ifStatement = "IF EXISTS(SELECT 1 FROM dbo.{} {})".format(sheetname, whereStatement) + '\r'
    updateValues = dict(zip(setList,valueList))
    updateValues = str(updateValues).replace("'",'').replace('{','').replace('}','').replace(':',' =').replace('  ', ' ').replace('created_dttm = GETDATE(), ','').replace('created_by = 0, ','').replace('"',"'")
    updateValues = updateValues.split(',',whereCount + 1)[whereCount + 1]
    updateStatement = "UPDATE dbo.{} SET{}".format(sheetname, updateValues) + '\r'

    return ifStatement + updateStatement + whereStatement


#
# Creates File with needed lines
#
def writeFile(queryList):
    header = """------------------------------------------------------------
-- DB_Change_ID:    {}
-- DB_ASSET:        {}
-- DB_RPE_VER:      V3_12_02_02
------------------------------------------------------------
DECLARE @ErrorCount INT;
DECLARE @ShouldApply VARCHAR(5);
DECLARE @ScriptID VARCHAR(30);
DECLARE @NA VARCHAR(16)
DECLARE @RPVer VARCHAR(16);
DECLARE @Asset VARCHAR(30);
SET @NA = 'NA'

SET @ScriptID = '{}';
SET @Asset = '{}';
SET @RPVer = 'V3_12_02_02';

EXEC Usp_geterrorcount @ScriptID, @ErrorCount OUTPUT

IF (@ErrorCount = 0)
BEGIN
  EXEC usp_ShouldApplyScript @ScriptID, @ShouldApply OUTPUT
  IF (@ShouldApply='TRUE')
    BEGIN TRY
      BEGIN

------------------------------------------------------------
-- This is where your script goes
""".format(scriptName, functionalArea, scriptName, functionalArea)

    footer = """
-- End of your script
------------------------------------------------------------

        EXEC usp_HandleScriptApplied @ScriptID, @Asset, @NA, @NA, @NA, @RPVer, @NA, @NA;
      END
    END TRY
    BEGIN CATCH
      EXEC usp_HandleErrorApplyingScript @ScriptID,@Asset,@NA,@NA,@NA,@RPVer,@NA,@NA;
    END CATCH;
  ELSE
    BEGIN
      PRINT 'Skipping previously applied script: '+@ScriptID+'. Please refer DB_CHANGE_ID in DATABASE_UPDATES table.';
    END
END
ELSE
      PRINT 'Skipping script: '+@ScriptID+'. Previous error detected.';
GO
"""

    completePath = "{}{}.sql".format(fileLocation,scriptName)

    fp = open(completePath, "w")
    fp.write(header + '\n')
    for i in queryList:
        fp.write(i + '\n\n')
    fp.write(footer)
    fp.close()

def main():
    workbook = openpyxl.load_workbook('{}{}.xlsx'.format(fileLocation,excelFileName))

    queryList = []

    for sheet in workbook:
        Sheet1 = workbook[sheet.title]
        numOfColumns = Sheet1.max_column
        numOfRows = Sheet1.max_row
        sheetname = sheet.title
        sheet1Data = Sheet1.iter_rows(min_row=0,max_row=numOfRows,min_col=0,max_col=numOfColumns,values_only=True)
        myList =  setUpList(sheet1Data)
        rowNames = createRowNames(myList[0])
        myList = formatValues(myList)

        for i in range(1, len(myList)):
            queryList.append(createQuery(myList[i], rowNames, sheetname))

    #for i in queryList:
    #    print(i)

        writeFile(queryList)


main()

